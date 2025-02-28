from flask import Blueprint, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from db_models import Inspection_form, Inspection_ticket
from io import BytesIO

exel_bp = Blueprint('excel', __name__)

# Обновленный словарь замен ключей
QUESTION_MAPPING = {
    "row1": "Степень огнестойкости и класс конструктивной пожарной опасности",
    "row2": "Состояние систем противопожарной защиты",
    "row3": "Состояние путей эвакуации",
    "row4": "Эксплуатация электрических сетей и оборудования",
    "row5": "Содержание территории и зданий",
    "row6": "Наличие первичных средств пожаротушения",
    "row7": "Соблюдение требований к вентиляционным системам",
    "row8": "Характеристика хранимых веществ и материалов",
    "row9": "Обеспечение противопожарного водоснабжения"
}

def get_excel(form_id):
    # Получаем запись из БД
    form = Inspection_form.query.get(form_id)
    if not form:
        return jsonify({"error": "Форма не найдена"}), 404

    # Получаем запись из Inspection_ticket
    ticket = Inspection_ticket.query.filter_by(assigment_id=form.assigment_id).first()
    
    # Преобразуем JSON-ответы, отфильтровываем и заменяем ключи
    json_data = form.answers
    data = [
        {"№": i + 1, "вопрос": QUESTION_MAPPING[q], "ответ": a}
        for i, (q, a) in enumerate(json_data.items())
        if q in QUESTION_MAPPING
    ]

    # Получаем значение conclusion
    conclusion = json_data.get("conclusion", "Нет данных")

    # Создаём Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = f"Inspection Report {form_id}"

    # Стили
    bold_center = Font(bold=True)
    center_align = Alignment(horizontal="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # Заголовки для основной таблицы
    headers = ["№", "Вопрос", "Ответ"]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = bold_center
        cell.alignment = center_align
        cell.border = thin_border

    # Заполняем основную таблицу
    for row_index, row in enumerate(data, start=2):
        ws.append([row["№"], row["вопрос"], row["ответ"]])
        for col_index in range(1, 4):
            ws.cell(row=row_index, column=col_index).border = thin_border

    # Вставляем conclusion ВПРАВО (в колонку E)
    conclusion_col = 5  # Колонка E
    ws.cell(row=1, column=conclusion_col, value="Заключение").font = bold_center
    ws.cell(row=1, column=conclusion_col).alignment = center_align
    ws.cell(row=1, column=conclusion_col).border = thin_border  # Границы для E1
    ws.cell(row=2, column=conclusion_col, value=conclusion).border = thin_border

    # **Добавляем таблицу Inspection_ticket справа (в колонках G-H)**
    if ticket:
        ticket_start_col = 7  # G-колонка
        ticket_start_row = 1

        # Заголовок таблицы в G1-H1
        ws.merge_cells(start_row=ticket_start_row, start_column=ticket_start_col, end_row=ticket_start_row, end_column=ticket_start_col + 1)
        ticket_header_cell = ws.cell(row=ticket_start_row, column=ticket_start_col, value="Данные о проверке")
        ticket_header_cell.font = bold_center
        ticket_header_cell.alignment = center_align
        ticket_header_cell.border = thin_border

        # Данные Inspection_ticket
        ticket_data = [
            ["Дата проверки", ticket.inspection_date.strftime("%d.%m.%Y") if ticket.inspection_date else "Нет данных"],
            ["Категория жителей", ticket.resident_category or "Нет данных"],
            ["Количество человек в семье", str(ticket.family_size) if ticket.family_size is not None else "Нет данных"]
        ]

        for row_index, row in enumerate(ticket_data, start=ticket_start_row + 1):
            for col_index, value in enumerate(row, start=ticket_start_col):
                cell = ws.cell(row=row_index, column=col_index, value=value)
                cell.border = thin_border  # Границы для всех ячеек таблицы

    # Автоширина колонок
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["E"].width = 30  # Ширина колонки заключения
    ws.column_dimensions["G"].width = 25  # Первая колонка Inspection_ticket
    ws.column_dimensions["H"].width = 20  # Вторая колонка Inspection_ticket

    # Сохраняем в память (не создавая файл на диске)
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(file_stream, as_attachment=True, download_name=f"inspection_report_{form_id}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


exel_bp.add_url_rule('/get_excel/<int:form_id>', view_func=get_excel, methods=['GET'])
